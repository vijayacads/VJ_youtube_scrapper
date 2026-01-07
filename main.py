import os
import csv
import io
import json
import uuid
import asyncio
import shutil
import zipfile
from pathlib import Path
from typing import List, Optional, Callable, Dict
from fastapi import FastAPI, HTTPException, UploadFile, File, Query, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, Response
from dotenv import load_dotenv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import openpyxl.utils
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from models import (
    YoutubeDetailsRequest, YoutubeDetailsResponse, YoutubeVideoFull, YoutubeError,
    YoutubeBulkRequest, ChannelExportRequest, ChannelExportResponse, JobStatus,
    ScraperRequest, ScraperResponse
)
from youtube_id import extract_video_id
from youtube_metadata import fetch_youtube_metadata
from youtube_transcript import fetch_transcript_text
from youtube_channel import fetch_channel_video_ids, resolve_channel_id, get_channel_title, get_channel_video_count, scrape_popular_videos

# Load environment variables in development
load_dotenv()

app = FastAPI(title="YouTube Scraper API", version="1.0.0")

# Simple job storage
jobs: Dict[str, Dict] = {}

# Serve static files (UI)
try:
    app.mount("/static", StaticFiles(directory="static"), name="static")
except:
    pass

@app.get("/")
async def root():
    """Serve the UI."""
    try:
        return FileResponse("static/index.html")
    except:
        return {"message": "UI not found. API is available at /docs"}


@app.get("/health")
async def health():
    """Health check endpoint."""
    return {"status": "ok"}


async def get_youtube_details(
    inputs: List[str], 
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    include_transcripts: bool = True
) -> YoutubeDetailsResponse:
    """
    Orchestrates fetching YouTube video details including metadata and transcripts.
    
    Steps:
    1. Extract video IDs from all inputs (handle errors for invalid inputs)
    2. Deduplicate IDs
    3. Fetch metadata in batches (exactly 50 IDs max per YouTube API call)
    4. For each video, fetch transcript (using blocking calls for current scale)
    5. Combine results into response with items and errors
    """
    items = []
    errors = []
    valid_ids = []
    id_to_input = {}  # Map video_id back to original input for error reporting
    
    # Step 1: Extract video IDs from all inputs
    for input_str in inputs:
        video_id = extract_video_id(input_str)
        if video_id is None:
            errors.append(YoutubeError(
                id_or_url=input_str,
                message="Invalid YouTube URL or ID format"
            ))
        else:
            # Track which input this ID came from
            if video_id not in id_to_input:
                id_to_input[video_id] = input_str
                valid_ids.append(video_id)
    
    # Step 2: Deduplicate IDs (already handled above, but ensure uniqueness)
    unique_ids = list(set(valid_ids))
    
    if not unique_ids:
        return YoutubeDetailsResponse(items=[], errors=errors)
    
    # Step 3: Fetch ALL metadata first (still in batches for efficiency)
    if progress_callback:
        progress_callback(0, len(unique_ids), "Fetching metadata for all videos...")
    
    metadata_dict = await fetch_youtube_metadata(unique_ids)
    
    # Update total based on actual videos found
    actual_total = len(metadata_dict)
    if not actual_total:
        return YoutubeDetailsResponse(items=[], errors=errors)
    
    if include_transcripts:
        if progress_callback:
            progress_callback(0, actual_total, f"Metadata fetched for {actual_total} videos. Processing transcripts...")
    else:
        if progress_callback:
            progress_callback(0, actual_total, f"Metadata fetched for {actual_total} videos. Skipping transcripts...")
    
    # Step 4: Process videos ONE BY ONE and report progress after EACH video
    seen_video_ids = set()  # Track processed video IDs to prevent duplicates
    processed_count = 0
    
    # Process in the order of unique_ids to maintain order
    for video_id in unique_ids:
        if video_id not in metadata_dict:
            continue  # Skip if metadata wasn't found
        
        if video_id in seen_video_ids:
            continue
        seen_video_ids.add(video_id)
        
        video = metadata_dict[video_id]
        
        # Fetch transcript for this video (only if requested)
        if include_transcripts:
            try:
                transcript = fetch_transcript_text(video_id)
                # Check if it was blocked vs just unavailable
                if transcript == "__BLOCKED__":
                    # Was blocked - already logged in fetch_transcript_text
                    video.transcript = None
                elif transcript is None:
                    # No transcript available (not blocked, just no captions)
                    video.transcript = None
                    print(f"ℹ️ No transcript available for {video_id} (video may not have captions)")
                else:
                    # Successfully fetched transcript
                    video.transcript = transcript
            except Exception as e:
                # If transcript fetch fails with unexpected error, set to None and continue
                video.transcript = None
                error_type = type(e).__name__
                if error_type in ['IpBlocked', 'RequestBlocked']:
                    print(f"⚠️ BLOCKED: Could not fetch transcript for {video_id} - YouTube is blocking requests")
                else:
                    print(f"⚠️ Warning: Could not fetch transcript for {video_id}: {error_type} - {str(e)[:100]}")
        else:
            # Transcripts disabled - set to None
            video.transcript = None
        
        # Add to results
        items.append(video)
        processed_count += 1
        
        # Report progress AFTER EACH video is fully processed
        if progress_callback:
            # #region agent log
            import json
            try:
                with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:157","message":"progress_callback called in get_youtube_details","data":{"current":processed_count,"total":actual_total,"video_id":video_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"A"}) + '\n')
            except: pass
            # #endregion
            # Truncate title for display if too long
            title_preview = video.title[:50] + "..." if len(video.title) > 50 else video.title
            progress_callback(
                processed_count, 
                actual_total, 
                f"Processed video {processed_count}/{actual_total}: {title_preview}"
            )
    
    # Step 5: Handle errors for IDs that weren't found in metadata
    found_ids = set(metadata_dict.keys())
    for video_id in unique_ids:
        if video_id not in found_ids:
            errors.append(YoutubeError(
                id_or_url=id_to_input.get(video_id, video_id),
                message="Video not found or metadata unavailable"
            ))
    
    return YoutubeDetailsResponse(items=items, errors=errors)


def log_progress(current: int, total: int, message: str):
    """Simple progress logging to console."""
    if total > 0:
        percentage = (current / total) * 100
        print(f"[Progress] {current}/{total} ({percentage:.1f}%) - {message}")
    else:
        print(f"[Progress] {message}")


@app.post("/youtube/details", response_model=YoutubeDetailsResponse)
async def youtube_details(request: YoutubeDetailsRequest):
    """
    Main endpoint to fetch YouTube video details including metadata and transcripts.
    
    Accepts either URLs or IDs (or both) and returns complete video information.
    Progress is logged to console automatically.
    """
    # Combine urls and ids into single list
    inputs = []
    if request.urls:
        inputs.extend(request.urls)
    if request.ids:
        inputs.extend(request.ids)
    
    # Validate input
    if not inputs:
        raise HTTPException(
            status_code=400,
            detail="At least one URL or ID must be provided"
        )
    
    # Get YouTube details with progress logging
    response = await get_youtube_details(inputs, progress_callback=log_progress)
    return response


def parse_bulk_input(content: str, content_type: str = "text/plain") -> List[str]:
    """
    Parses bulk input from various formats.
    Returns list of URLs/IDs.
    """
    urls = []
    
    if content_type.startswith("text/plain") or content_type == "text/csv":
        # Plain text or CSV - split by newlines
        lines = content.strip().split('\n')
        for line in lines:
            line = line.strip()
            if line and not line.startswith('#'):  # Skip empty lines and comments
                # For CSV, take first column
                if ',' in line:
                    line = line.split(',')[0].strip()
                # Remove quotes if present
                line = line.strip('"').strip("'")
                if line:
                    urls.append(line)
    
    elif content_type == "application/json":
        # JSON array
        import json
        try:
            data = json.loads(content)
            if isinstance(data, list):
                urls = [str(item) for item in data if item]
        except:
            pass
    
    return urls


@app.post("/youtube/details/bulk")
async def youtube_details_bulk(
    http_request: Request,
    file: Optional[UploadFile] = File(None)
):
    """
    Bulk endpoint to fetch YouTube video details.
    
    Accepts either:
    - JSON body with 'urls_text' field (plain text, newline-separated URLs)
    - File upload (TXT, CSV, or JSON) via multipart/form-data
    
    Returns a job_id immediately. Check progress and get result via /jobs/{job_id}
    """
    inputs = []
    
    # Check content type to determine how to parse
    content_type = http_request.headers.get("content-type", "").lower()
    
    # Process file upload if provided (multipart/form-data)
    if file:
        content = await file.read()
        content_str = content.decode('utf-8')
        file_content_type = file.content_type or "text/plain"
        inputs.extend(parse_bulk_input(content_str, file_content_type))
    # Process JSON body if provided (application/json)
    include_transcripts = True  # Default to True for backward compatibility
    if "application/json" in content_type:
        try:
            body = await http_request.json()
            if "urls_text" in body and body["urls_text"]:
                inputs.extend(parse_bulk_input(body["urls_text"], "text/plain"))
            # Extract include_transcripts from request body
            if "include_transcripts" in body:
                include_transcripts = bool(body["include_transcripts"])
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid JSON body: {str(e)}"
            )
    
    # Validate input
    if not inputs:
        raise HTTPException(
            status_code=400,
            detail="Either a file or urls_text in request body must be provided"
        )
    
    job_id = str(uuid.uuid4())
    cancellation_token = asyncio.Event()
    
    # Initialize job
    jobs[job_id] = {
        "status": "running",
        "current": 0,
        "total": len(inputs),
        "message": "Starting...",
        "cancellation_token": cancellation_token,
        "result": None
    }
    
    # Start processing in background
    asyncio.create_task(process_bulk_details(job_id, inputs, cancellation_token, include_transcripts))
    
    return {"job_id": job_id, "message": "Job started. Check /jobs/{job_id} for progress."}


async def process_bulk_details(job_id: str, inputs: List[str], cancellation_token: asyncio.Event, include_transcripts: bool = True):
    """Background task to process bulk video details."""
    try:
        # Progress callback - ensure updates are immediately visible
        def progress_callback(current: int, total: int, message: str):
            # #region agent log
            import json
            try:
                with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:315","message":"progress_callback entry in process_bulk_details","data":{"current":current,"total":total,"job_id":job_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"A"}) + '\n')
            except: pass
            # #endregion
            if not cancellation_token.is_set() and job_id in jobs:
                # Update job status - create new dict to ensure visibility
                jobs[job_id].update({
                    "current": current,
                    "total": total,
                    "message": message,
                    "status": "running"
                })
                # #region agent log
                try:
                    with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:318","message":"job dict updated (before reassign)","data":{"current":jobs[job_id]["current"],"total":jobs[job_id]["total"],"job_id":job_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"A"}) + '\n')
                except: pass
                # #endregion
                # Force dict update by reassigning (helps with visibility)
                jobs[job_id] = dict(jobs[job_id])
                # #region agent log
                try:
                    with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:325","message":"job dict reassigned (after reassign)","data":{"current":jobs[job_id]["current"],"total":jobs[job_id]["total"],"job_id":job_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"A"}) + '\n')
                except: pass
                # #endregion
        
        # Get YouTube details
        response = await get_youtube_details(inputs, progress_callback=progress_callback, include_transcripts=include_transcripts)
        
        if cancellation_token.is_set():
            jobs[job_id]["status"] = "cancelled"
            return
        
        jobs[job_id] = {
            "status": "completed",
            "current": len(response.items),
            "total": len(response.items),
            "message": "Completed",
            "result": response.dict()
        }
        
    except asyncio.CancelledError:
        jobs[job_id]["status"] = "cancelled"
        jobs[job_id]["message"] = "Cancelled by user"
    except Exception as e:
        jobs[job_id] = {
            "status": "error",
            "current": 0,
            "total": 0,
            "message": f"Error: {str(e)}",
            "result": None
        }


def convert_to_csv(response: YoutubeDetailsResponse) -> str:
    """
    Converts YoutubeDetailsResponse to CSV format.
    Properly handles special characters, newlines, and long text fields.
    Uses QUOTE_MINIMAL (default) which automatically quotes fields containing special chars.
    """
    output = io.StringIO()
    # Use default quoting (QUOTE_MINIMAL) which automatically quotes fields with commas, quotes, or newlines
    # This is the safest approach for CSV compatibility
    writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
    
    # Write header
    writer.writerow([
        "id", "url", "title", "description", "channel_title",
        "published_at", "duration", "transcript"
    ])
    
    # Track seen IDs to prevent duplicates
    seen_ids = set()
    
    # Write data rows
    for item in response.items:
        # Skip duplicates (shouldn't happen, but safety check)
        if item.id in seen_ids:
            print(f"Warning: Duplicate video ID found: {item.id}, skipping...")
            continue
        seen_ids.add(item.id)
        
        # Ensure all fields are strings and handle None values
        transcript = str(item.transcript) if item.transcript else ""
        
        writer.writerow([
            str(item.id) if item.id else "",
            str(item.url) if item.url else "",
            str(item.title) if item.title else "",
            str(item.description) if item.description else "",
            str(item.channel_title) if item.channel_title else "",
            str(item.published_at) if item.published_at else "",
            str(item.duration) if item.duration else "",
            transcript
        ])
    
    return output.getvalue()


def convert_to_excel(response: YoutubeDetailsResponse) -> bytes:
    """
    Converts YoutubeDetailsResponse to Excel format.
    Each field is in one column, all data in rows.
    """
    if not OPENPYXL_AVAILABLE:
        raise ValueError("openpyxl is not installed. Install it with: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube Videos"
    
    # Header row with styling
    headers = ["ID", "URL", "Title", "Description", "Channel Title", 
               "Published At", "Duration", "Transcript"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c4869", end_color="2c4869", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_num, item in enumerate(response.items, 2):
        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=item.url)
        ws.cell(row=row_num, column=3, value=item.title)
        ws.cell(row=row_num, column=4, value=item.description)
        ws.cell(row=row_num, column=5, value=item.channel_title)
        ws.cell(row=row_num, column=6, value=item.published_at)
        ws.cell(row=row_num, column=7, value=item.duration)
        ws.cell(row=row_num, column=8, value=item.transcript or "")
        
        # Set alignment for transcript column (wrap text)
        transcript_cell = ws.cell(row=row_num, column=8)
        transcript_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths
    column_widths = [15, 40, 40, 60, 25, 20, 15, 80]
    for col_num, width in enumerate(column_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@app.post("/youtube/channel/export")
async def channel_export(request: ChannelExportRequest):
    """
    Export all videos from a YouTube channel.
    
    Returns a job_id immediately. Check progress and get result via /jobs/{job_id}
    """
    job_id = str(uuid.uuid4())
    cancellation_token = asyncio.Event()
    
    # Initialize job
    jobs[job_id] = {
        "status": "running",
        "current": 0,
        "total": 0,
        "message": "Starting...",
        "cancellation_token": cancellation_token,
        "result": None
    }
    
    # Start processing in background
    asyncio.create_task(process_channel_export(job_id, request, cancellation_token))
    
    return {"job_id": job_id, "message": "Job started. Check /jobs/{job_id} for progress."}


async def process_channel_export(job_id: str, request: ChannelExportRequest, cancellation_token: asyncio.Event):
    """Background task to process channel export."""
    try:
        # Resolve channel ID
        resolved_id = await resolve_channel_id(request.channel_id_or_url)
        if not resolved_id:
            jobs[job_id] = {
                "status": "error",
                "current": 0,
                "total": 0,
                "message": "Invalid channel ID or URL",
                "result": None
            }
            return
        
        # Check channel video count
        jobs[job_id]["message"] = "Checking channel size..."
        video_count = await get_channel_video_count(resolved_id)
        if video_count is not None and video_count > 500:
            jobs[job_id] = {
                "status": "error",
                "current": 0,
                "total": 0,
                "message": f"Channel has {video_count} videos (limit: 500). For large channels, use the Scraper tab to get top videos, then paste URLs into Bulk Videos tab.",
                "result": None
            }
            return
        
        jobs[job_id]["message"] = "Fetching video list..."
        
        # Fetch video IDs from channel (with sort_by parameter)
        video_ids = await fetch_channel_video_ids(resolved_id, request.max_videos, request.sort_by)
        
        if cancellation_token.is_set():
            jobs[job_id]["status"] = "cancelled"
            return
        
        if not video_ids:
            jobs[job_id] = {
                "status": "error",
                "current": 0,
                "total": 0,
                "message": "No videos found for this channel",
                "result": None
            }
            return
        
        # Get channel title
        channel_title = await get_channel_title(resolved_id)
        
        jobs[job_id]["total"] = len(video_ids)
        jobs[job_id]["message"] = f"Found {len(video_ids)} videos. Processing..."
        
        # Progress callback
        def progress_callback(current: int, total: int, message: str):
            # #region agent log
            import json
            try:
                with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                    f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:520","message":"progress_callback entry in process_channel_export","data":{"current":current,"total":total,"job_id":job_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"D"}) + '\n')
            except: pass
            # #endregion
            if not cancellation_token.is_set() and job_id in jobs:
                jobs[job_id]["current"] = current
                jobs[job_id]["total"] = total
                jobs[job_id]["message"] = f"Processing video {current}/{total}: {message}"
                # #region agent log
                try:
                    with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
                        f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:523","message":"job dict updated in process_channel_export","data":{"current":jobs[job_id]["current"],"total":jobs[job_id]["total"],"job_id":job_id},"sessionId":"debug-session","runId":"run1","hypothesisId":"D"}) + '\n')
                except: pass
                # #endregion
        
        # If sorting by popular, we need to fetch metadata first, sort by view_count, then process
        if request.sort_by == "popular":
            jobs[job_id]["message"] = "Fetching metadata for sorting..."
            # Fetch metadata with statistics
            metadata_dict = await fetch_youtube_metadata(video_ids)
            
            if cancellation_token.is_set():
                jobs[job_id]["status"] = "cancelled"
                return
            
            # Sort by view_count (descending), handling None values
            sorted_videos = sorted(
                metadata_dict.values(),
                key=lambda v: v.view_count if v.view_count is not None else 0,
                reverse=True
            )
            
            # Take top max_videos (or all if None)
            if request.max_videos:
                sorted_videos = sorted_videos[:request.max_videos]
            
            # Extract video IDs in sorted order
            video_ids = [v.id for v in sorted_videos]
            jobs[job_id]["total"] = len(video_ids)
            jobs[job_id]["message"] = f"Sorted {len(video_ids)} videos by popularity. Processing..."
        
        # Fetch details for all videos
        details_response = await get_youtube_details(
            [f"https://www.youtube.com/watch?v={vid}" for vid in video_ids],
            progress_callback=progress_callback
        )
        
        if cancellation_token.is_set():
            jobs[job_id]["status"] = "cancelled"
            return
        
        # Optionally skip transcripts if not requested
        if not request.include_transcripts:
            print(f"⚠️ Transcripts disabled by user request for channel export")
            for item in details_response.items:
                item.transcript = None
        else:
            # Log transcript status for debugging
            transcripts_found = sum(1 for item in details_response.items if item.transcript is not None)
            transcripts_missing = len(details_response.items) - transcripts_found
            print(f"ℹ️ Channel export complete: {transcripts_found} videos with transcripts, {transcripts_missing} without transcripts")
        
        # Always store as JSON/YoutubeDetailsResponse format
        # Format conversion happens at download time based on user's choice
        result = ChannelExportResponse(
            channel_id=resolved_id,
            channel_title=channel_title,
            total_videos=len(video_ids),
            processed_videos=len(details_response.items),
            data=details_response,  # Always store as YoutubeDetailsResponse
            errors=details_response.errors
        )
        
        jobs[job_id] = {
            "status": "completed",
            "current": len(details_response.items),
            "total": len(video_ids),
            "message": "Completed",
            "result": result.model_dump() if hasattr(result, 'model_dump') else result.dict()
        }
        
    except asyncio.CancelledError:
        jobs[job_id]["status"] = "cancelled"
        jobs[job_id]["message"] = "Cancelled by user"
    except Exception as e:
        jobs[job_id] = {
            "status": "error",
            "current": 0,
            "total": 0,
            "message": f"Error: {str(e)}",
            "result": None
        }


@app.post("/youtube/scraper/extract")
async def scraper_extract(request: ScraperRequest):
    """
    Extract video URLs from YouTube channel's Popular page using web scraping.
    
    Returns a job_id immediately. Check progress and get result via /jobs/{job_id}
    """
    job_id = str(uuid.uuid4())
    cancellation_token = asyncio.Event()
    
    # Initialize job
    jobs[job_id] = {
        "status": "running",
        "current": 0,
        "total": 0,
        "message": "Starting...",
        "cancellation_token": cancellation_token,
        "result": None
    }
    
    # Start processing in background
    asyncio.create_task(process_scraper_extract(job_id, request, cancellation_token))
    
    return {"job_id": job_id, "message": "Job started. Check /jobs/{job_id} for progress."}


async def process_scraper_extract(job_id: str, request: ScraperRequest, cancellation_token: asyncio.Event):
    """Background task to process scraper extraction."""
    try:
        content_type_label = "shorts" if request.content_type == "shorts" else "videos"
        jobs[job_id]["message"] = f"Scraping Popular {content_type_label} page..."
        
        # Scrape popular videos or shorts
        video_urls = await scrape_popular_videos(request.channel_url, request.max_videos, request.content_type)
        
        if cancellation_token.is_set():
            jobs[job_id]["status"] = "cancelled"
            return
        
        if not video_urls:
            jobs[job_id] = {
                "status": "error",
                "current": 0,
                "total": 0,
                "message": "No videos found. Please check the channel URL.",
                "result": None
            }
            return
        
        # Create response
        result = ScraperResponse(
            video_urls=video_urls,
            count=len(video_urls)
        )
        
        jobs[job_id] = {
            "status": "completed",
            "current": len(video_urls),
            "total": len(video_urls),
            "message": f"Extracted {len(video_urls)} video URLs",
            "result": result.model_dump() if hasattr(result, 'model_dump') else result.dict()
        }
        
    except ImportError as e:
        jobs[job_id] = {
            "status": "error",
            "current": 0,
            "total": 0,
            "message": f"Playwright not installed: {str(e)}. Install with: pip install playwright && playwright install chromium",
            "result": None
        }
    except Exception as e:
        jobs[job_id] = {
            "status": "error",
            "current": 0,
            "total": 0,
            "message": f"Error scraping videos: {str(e)}",
            "result": None
        }


@app.get("/jobs/{job_id}", response_model=JobStatus)
async def get_job_status(job_id: str):
    """
    Get job status and progress.
    Shows current/total progress and allows cancellation.
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    # #region agent log
    import json
    try:
        with open(r'c:\Users\vijay\Documents\GitHub\VJ_youtube_scrapper\.cursor\debug.log', 'a', encoding='utf-8') as f:
            f.write(json.dumps({"id":f"log_{int(__import__('time').time()*1000)}","timestamp":int(__import__('time').time()*1000),"location":"main.py:582","message":"get_job_status endpoint called","data":{"job_id":job_id,"current":job["current"],"total":job["total"],"status":job["status"]},"sessionId":"debug-session","runId":"run1","hypothesisId":"A"}) + '\n')
    except: pass
    # #endregion
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        current=job["current"],
        total=job["total"],
        message=job["message"],
        result=job.get("result")
    )


@app.post("/jobs/{job_id}/cancel")
async def cancel_job(job_id: str):
    """
    Cancel a running job.
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job["status"] not in ["running"]:
        raise HTTPException(status_code=400, detail=f"Job is {job['status']}, cannot cancel")
    
    job["cancellation_token"].set()
    job["status"] = "cancelling"
    job["message"] = "Cancellation requested..."
    
    return {"job_id": job_id, "status": "cancelling", "message": "Cancellation requested"}


@app.get("/jobs/{job_id}/download")
async def download_job_result(job_id: str, format: str = Query("json", pattern="^(json|csv|excel)$")):
    """
    Download job result in specified format (json, csv, excel).
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = jobs[job_id]
    if job.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Job not completed yet")
    
    result = job.get("result")
    if not result:
        raise HTTPException(status_code=404, detail="No result available")
    
    format_lower = format.lower()
    
    # Handle ChannelExportResponse - extract the YoutubeDetailsResponse from it
    if isinstance(result, dict) and "channel_id" in result:
        from models import ChannelExportResponse, YoutubeDetailsResponse
        
        # Parse the ChannelExportResponse
        export_resp = ChannelExportResponse(**result)
        
        # Extract the YoutubeDetailsResponse from the data field
        if isinstance(export_resp.data, dict):
            # Data is a dict (YoutubeDetailsResponse)
            details = YoutubeDetailsResponse(**export_resp.data)
        elif isinstance(export_resp.data, str):
            # Old format: data was stored as CSV string - can't convert back, return error
            raise HTTPException(status_code=400, detail="Data stored in old format. Please re-export.")
        else:
            # Should be YoutubeDetailsResponse object
            details = export_resp.data
        
        # Now convert based on requested format
        if format_lower == "excel":
            excel_data = convert_to_excel(details)
            return Response(
                content=excel_data,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="channel-export-{job_id}.xlsx"'}
            )
        elif format_lower == "csv":
            csv_data = convert_to_csv(details)
            return Response(
                content=csv_data,
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="channel-export-{job_id}.csv"'}
            )
        else:
            # JSON format - return the full ChannelExportResponse
            return Response(
                content=json.dumps(result, indent=2),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="channel-export-{job_id}.json"'}
            )
    
    # Handle YoutubeDetailsResponse
    if isinstance(result, dict) and "items" in result:
        from models import YoutubeDetailsResponse
        response = YoutubeDetailsResponse(**result)
        
        if format_lower == "excel":
            excel_data = convert_to_excel(response)
            return Response(
                content=excel_data,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="youtube-export-{job_id}.xlsx"'}
            )
        elif format_lower == "csv":
            csv_data = convert_to_csv(response)
            return Response(
                content=csv_data,
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="youtube-export-{job_id}.csv"'}
            )
        else:
            # JSON format
            return Response(
                content=json.dumps(result, indent=2),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="youtube-export-{job_id}.json"'}
            )
    
    # Fallback to JSON
    return Response(
        content=json.dumps(result, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="export-{job_id}.json"'}
    )


@app.get("/download-package")
async def download_package():
    """
    Download the standalone package as a zip file.
    Creates a distributable package on-the-fly.
    """
    try:
        # Create temporary package directory
        package_name = "youtube-scraper-standalone"
        temp_dir = Path("temp_package")
        temp_dir.mkdir(exist_ok=True)
        package_dir = temp_dir / package_name
        
        # Clean up if exists
        if package_dir.exists():
            shutil.rmtree(package_dir)
        package_dir.mkdir(parents=True, exist_ok=True)
        
        # Files to include
        files_to_include = [
            "main.py",
            "models.py",
            "youtube_id.py",
            "youtube_metadata.py",
            "youtube_transcript.py",
            "youtube_channel.py",
            "requirements.txt",
            "install.bat",
            "install.sh",
            "run.bat",
            "run.sh",
            "README_PACKAGE.md",
        ]
        
        # Copy files
        for file in files_to_include:
            if os.path.exists(file):
                shutil.copy2(file, package_dir / file)
        
        # Copy static directory
        if os.path.exists("static"):
            shutil.copytree("static", package_dir / "static", dirs_exist_ok=True)
        
        # Create zip file in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(package_dir):
                # Skip __pycache__ and .pyc files
                dirs[:] = [d for d in dirs if d != '__pycache__']
                files = [f for f in files if not f.endswith('.pyc')]
                
                for file in files:
                    file_path = Path(root) / file
                    arcname = file_path.relative_to(package_dir)
                    zipf.write(file_path, arcname)
        
        # Clean up temp directory
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        # Return zip file
        zip_buffer.seek(0)
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": "attachment; filename=youtube-scraper-standalone.zip"
            }
        )
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating package: {str(e)}")


@app.post("/shutdown")
async def shutdown_server():
    """
    Shutdown the server gracefully.
    Only works in standalone mode (not on Render).
    """
    import sys
    
    # Check if running on Render (has RENDER env var)
    if os.getenv("RENDER"):
        raise HTTPException(
            status_code=403, 
            detail="Shutdown is not available on Render. This endpoint only works in standalone mode."
        )
    
    # For standalone mode, shutdown the server
    async def shutdown_background():
        await asyncio.sleep(0.5)  # Give time for response to be sent
        os._exit(0)
    
    # Start background task to shutdown
    asyncio.create_task(shutdown_background())
    
    return {
        "message": "Server is shutting down...",
        "status": "shutting_down"
    }


if __name__ == "__main__":
    import uvicorn
    print("=" * 60)
    print("Starting YouTube Scraper Server...")
    print("=" * 60)
    print("Server will be available at: http://localhost:8000")
    print("Press Ctrl+C to stop the server")
    print("=" * 60)
    print()
    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")


